import json
import time
import math
import io
import random
import threading
import sys
import os
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from typing import List, Optional
from pathlib import Path

# 嘗試引入 GUI 視窗套件
try:
    import webview
except ImportError:
    print("[ERROR] Pywebview not found.")
    sys.exit(1)

# 嘗試引入其他套件
try:
    from geopy.geocoders import Nominatim
except ImportError:
    Nominatim = None
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    openpyxl = None

app = FastAPI(title="Tohoku Trip App")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- 路徑判斷邏輯 ---
if getattr(sys, 'frozen', False):
    # EXE/App 模式
    BUNDLE_DIR = Path(sys._MEIPASS)
    FRONTEND_DIR = BUNDLE_DIR / "frontend" / "out"
    # 🔥 修正：單一執行檔模式，db.json 應該在執行檔旁邊
    if sys.argv[0].endswith('.app/Contents/MacOS/TohokuTrip'):
        # Mac App Bundle 模式
        APP_DIR = Path(sys.executable).parent.parent / "MacOS"
    else:
        # 單一執行檔模式
        APP_DIR = Path(sys.executable).parent
else:
    # 開發模式
    BASE_DIR = Path(__file__).resolve().parent
    FRONTEND_DIR = BASE_DIR.parent / "frontend" / "out"
    APP_DIR = BASE_DIR

# 設定資料庫路徑
DB_FILE_PATH = APP_DIR / "db.json"

# 🔥 修改：移除 Emoji，改用純文字 Log，避免 Windows CP950 報錯
print(f"[INFO] App Path: {APP_DIR}")
print(f"[INFO] Frontend Path: {FRONTEND_DIR}")

# ... (資料模型與邏輯保持不變) ...
class Weather(BaseModel):
    icon: str; temp: str; desc: str
class Stop(BaseModel):
    time: str; name: str; description: Optional[str] = ""; transport: Optional[str] = ""; cost: int; currency: Optional[str] = "JPY"; lat: Optional[float] = 0; lng: Optional[float] = 0
class Day(BaseModel):
    dayNumber: int; date: str; theme: str; weather: Optional[Weather] = None; alternatives: Optional[str] = ""; checklist: Optional[List[str]] = []
    stops: List[Stop]; accommodation: Optional[str] = ""; accommodation_cost: Optional[int] = 0; accommodation_currency: Optional[str] = "JPY"
class TripMeta(BaseModel):
    title: str; days_count: int; travelers: int; budget: int; location: Optional[str] = "日本東北"; start_date: Optional[str] = "2026-04-15"
    home_currency: Optional[str] = "TWD"; destination_currency: Optional[str] = "JPY"; exchange_rate: Optional[float] = 0.215
class Itinerary(BaseModel):
    trip_meta: TripMeta; days: List[Day]

DEFAULT_ITINERARY = {
  "trip_meta": { "title": "2026 東北櫻花夢幻之旅", "days_count": 7, "travelers": 2, "budget": 150000, "location": "日本東北", "start_date": "2026-04-15", "home_currency": "TWD", "destination_currency": "JPY", "exchange_rate": 0.215 },
  "days": [ { "dayNumber": 1, "date": "2026-04-15 (週三)", "theme": "抵達仙台", "stops": [ { "time": "14:35", "name": "仙台機場 (SDJ)", "description": "抵達與入境", "transport": "", "cost": 0, "currency": "JPY" }, { "time": "16:30", "name": "仙台大都會飯店", "description": "Check-in", "transport": "仙台空港Access線", "cost": 660, "currency": "JPY" } ], "accommodation": "仙台大都會飯店", "accommodation_cost": 15000, "accommodation_currency": "JPY" } ]
}

@app.get("/api/itinerary")
def get_itinerary():
    if not DB_FILE_PATH.exists():
        try:
            with open(DB_FILE_PATH, "w", encoding="utf-8") as f:
                json.dump(DEFAULT_ITINERARY, f, indent=2, ensure_ascii=False)
        except PermissionError:
            print("[WARN] Permission denied for db.json")
            return DEFAULT_ITINERARY
        return DEFAULT_ITINERARY
    try:
        with open(DB_FILE_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except:
        return DEFAULT_ITINERARY

@app.post("/api/itinerary")
def update_itinerary(data: Itinerary):
    try:
        print("[INFO] Saving itinerary...")
        with open(DB_FILE_PATH, "w", encoding="utf-8") as f:
            json.dump(data.dict(), f, indent=2, ensure_ascii=False)
        return {"status": "success", "message": "Saved"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/export")
def export_report():
    if not openpyxl: raise HTTPException(status_code=500, detail="Missing openpyxl")
    if not DB_FILE_PATH.exists(): raise HTTPException(status_code=404, detail="No Data")
    
    with open(DB_FILE_PATH, "r", encoding="utf-8") as f: data = json.load(f)
    meta = data.get('trip_meta', {})
    home_curr = meta.get('home_currency', 'TWD')
    rate = meta.get('exchange_rate', 0.215)

    wb = openpyxl.Workbook()
    header_font = Font(name='微軟正黑體', bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_font = Font(name='微軟正黑體', bold=True, size=16, color="1F4E78")
    center_align = Alignment(horizontal="center", vertical="center")
    
    ws = wb.active; ws.title = "行程總覽"
    ws.append(["日期", "行程重點"])
    for day in data.get('days', []):
        ws.append([day['date'], day['theme']])

    output = io.BytesIO(); wb.save(output); output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=Trip_Itinerary.xlsx"})

# 掛載靜態檔案
if FRONTEND_DIR.exists():
    app.mount("/", StaticFiles(directory=str(FRONTEND_DIR), html=True), name="static")
else:
    print(f"[WARN] Frontend directory not found: {FRONTEND_DIR}")

# 啟動邏輯
def start_server():
    import uvicorn
    # log_level="critical" 減少輸出，避免編碼錯誤
    uvicorn.run(app, host="127.0.0.1", port=49152, log_level="critical")

if __name__ == "__main__":
    t = threading.Thread(target=start_server)
    t.daemon = True
    t.start()
    time.sleep(1)
    
    print("[INFO] Starting Window...")
    webview.create_window(
        "2026 東北櫻花夢幻之旅", 
        "http://127.0.0.1:49152",
        width=1280, 
        height=800, 
        resizable=True
    )
    webview.start()